"""
Filename: financial.py
Author: Tyler S. Budd
Current Status: In Development
"""
__author__ = "Tyler S. Budd"
__version__ = "1.000"

import openpyxl
import os
import time
import mysql.connector as mysql
from openpyxl.utils import column_index_from_string
from cellStyle import *
import sshtunnel

class ExcelSheet():

    def __init__(self):
        # change path
        # self.path = "D:/Users/Tyler/Documents/Github/financialSystem"
        # os.chdir (self.path)

        #import excel sheet
        self.filename = 'Example Spreadsheet.xlsx'
        self.wb = openpyxl.load_workbook(self.filename)

        #getting sheets from excel sheet
        self.home = self.wb['HOME']
        self.paychecks = self.wb['Paychecks']
        self.dataTrack = self.wb['databaseTrack']
        self.january = self.wb['January']
        self.febuary = self.wb['Febuary']
        self.march = self.wb['March']
        self.april = self.wb['April']
        self.may = self.wb['May']
        self.june = self.wb['June']
        self.july = self.wb['July']
        self.august = self.wb['August']
        self.september = self.wb['September']
        self.october = self.wb['October']
        self.november = self.wb['November']
        self.december = self.wb['December']

    def saveFile(self):
        sheet = self.wb
        file = self.filename
        sheet.save(file)

    def cellStyle(self, align, color, cell):
        cell.fill = color
        cell.border = Style.ALLBORDER
        cell.alignment = align

class DatabaseQuery():
    #initialzing database connection
    def __init__(self, db):
        self.db = db
        self.cursor = self.db.cursor()

    def closeConnection(self):
        self.db.close()

    def testDBQuery(self): # works
        cursor = self.cursor
        query = "select * from income"

        cursor.execute(query)
        result = cursor.fetchall()

        for x in result:
            print(x)

        query = 'call newIncome(350.00, "rent", "Mom", "2021-12-01", "Housing")'

        cursor.execute(query)
        
        self.db.commit()

    # ----------------------------------------------HOME sheet----------------------------------
    def updateSubscription(self):
        """
        Downloads Subscription info from database to spreadsheet
        Implemented
        """
        cursor = self.cursor

        wb = ExcelSheet()
        home = wb.home

        """
        getting all subscriptions from database
        """
        query = """select * from subscription"""
        cursor.execute(query)
        result = cursor.fetchall() #array of 5 (not using first column)

        """
        Clearing all old data
        """
        for i in range(18, 22):
            home.cell(row = i, column= column_index_from_string("A")).value = None
            home.cell(row = i, column= column_index_from_string("B")).value = None
            home.cell(row = i, column= column_index_from_string("C")).value = None
            home.cell(row = i, column= column_index_from_string("D")).value = None

        """
        downloading/updating all subscription data to spreadsheet
        """
        i = 18 # starting row [FIX FOR FUTURE USE]
        for tracked in result:
            item = tracked[1]
            amount = tracked[2]
            start = tracked[3]
            status = tracked[4]

            date = start.strftime("%m/%d/%Y")

            home.cell(row = i, column= column_index_from_string("A")).value = item
            home.cell(row = i, column= column_index_from_string("B")).value = date
            home.cell(row = i, column= column_index_from_string("C")).value = status
            home.cell(row = i, column= column_index_from_string("D")).value = amount

            wb.cellStyle(Style.LEFTALIGN, HomeColor.SUBSCRIPTIONFILL,
                        home.cell(row = i, column= column_index_from_string("A")))
            wb.cellStyle(Style.CENTERALIGN, HomeColor.SUBSCRIPTIONFILL,
                        home.cell(row = i, column= column_index_from_string("B")))
            wb.cellStyle(Style.CENTERALIGN, HomeColor.SUBSCRIPTIONFILL,
                        home.cell(row = i, column= column_index_from_string("C")))
            wb.cellStyle(Style.RIGHTALIGN, HomeColor.SUBSCRIPTIONFILL,
                        home.cell(row = i, column= column_index_from_string("D")))

            i += 1

            if i > 22:
                raise ValueError("\tSIZE ERROR: Number of subscriptions have exceeded the total allowed on %s\n" % wb.filename)
            
        self.closeConnection()
        wb.saveFile()

    def updateDesiredPurchase(self):
        """
        Downloads Desired Purchase info from database to spreadsheet"""
        cursor = self.cursor

        wb = ExcelSheet()
        home = wb.home

        """
        getting all desired purchases from database
        """
        query = """select * from desiredPur"""
        cursor.execute(query)
        result = cursor.fetchall() #array of 4 (not using first column)

        """
        Clearing all old data
        """
        for i in range(28, 32):
            home.cell(row = i, column= column_index_from_string("A")).value = None
            home.cell(row = i, column= column_index_from_string("B")).value = None
            home.cell(row = i, column= column_index_from_string("C")).value = None

        """
        downloading/updating all subscription data to spreadsheet
        """
        i = 28 # starting row [FIX FOR FUTURE USE]
        for tracked in result:
            item = tracked[1]
            amount = tracked[2]
            status = tracked[3]

            home.cell(row = i, column= column_index_from_string("A")).value = item
            home.cell(row = i, column= column_index_from_string("B")).value = status
            home.cell(row = i, column= column_index_from_string("C")).value = amount

            wb.cellStyle(Style.LEFTALIGN, HomeColor.DESIREDPURFILL,
                        home.cell(row = i, column= column_index_from_string("A")))
            wb.cellStyle(Style.CENTERALIGN, HomeColor.DESIREDPURFILL,
                        home.cell(row = i, column= column_index_from_string("B")))
            wb.cellStyle(Style.RIGHTALIGN, HomeColor.DESIREDPURFILL,
                        home.cell(row = i, column= column_index_from_string("C")))

            i += 1

            if i > 32:
                raise ValueError("\tSIZE ERROR: Number of desired purchases have exceeded the total allowed on %s\n" % wb.filename)
            
        self.closeConnection()
        wb.saveFile()

    def updateForSale(self):
        """
        Downloads For Sale info from database to spreadsheet
        Not Implemented
        """
        cursor = self.cursor

        wb = ExcelSheet()
        home = wb.home

        """
        getting all sales from database
        """
        query = """select * from forSale"""
        cursor.execute(query)
        result = cursor.fetchall() #array of 5 (not using first column)

        """
        Clearing all old data
        """
        for i in range(38, 40):
            home.cell(row = i, column= column_index_from_string("A")).value = None
            home.cell(row = i, column= column_index_from_string("B")).value = None
            home.cell(row = i, column= column_index_from_string("C")).value = None

        """
        downloading/updating all for sale data to spreadsheet
        """
        i = 38 # starting row [FIX FOR FUTURE USE]
        for tracked in result:
            item = tracked[1]
            amount = tracked[2]
            status = tracked[3]

            home.cell(row = i, column= column_index_from_string("A")).value = item
            home.cell(row = i, column= column_index_from_string("B")).value = status
            home.cell(row = i, column= column_index_from_string("C")).value = amount

            wb.cellStyle(Style.LEFTALIGN, HomeColor.FORSALEFILL,
                        home.cell(row = i, column= column_index_from_string("A")))
            wb.cellStyle(Style.CENTERALIGN, HomeColor.FORSALEFILL,
                        home.cell(row = i, column= column_index_from_string("B")))
            wb.cellStyle(Style.RIGHTALIGN, HomeColor.FORSALEFILL,
                        home.cell(row = i, column= column_index_from_string("C")))

            i += 1

            if i > 40:
                raise ValueError("\tSIZE ERROR: Number of for-sale items have exceeded the total allowed on %s\n" % wb.filename)
            
        self.closeConnection()
        wb.saveFile()

    def updateNetWorth(self):
        """
        Updates Net Worth Table in HOME spreadsheet
        Not Implemented
        """
        cursor = self.cursor

        wb = ExcelSheet()
        home = wb.home

        """
        Getting most recent net worth data from database
        """
        query = "call netWorth"
        cursor.execute(query)
        result = cursor.fetchall() #array of 3

        """
        downloading all account data to spreadsheet
        """
        for tracked in result:
            amount = tracked[0]
            account = tracked[1]
            day = tracked[2]

            date = day.strftime("%m/%d/%Y")

            # if account is bank account (total in bank is autocalculated in excel)
            if "Bank Account" in account:
                if "Spend" in account:
                    home.cell(row = 22, column= column_index_from_string("L")).value = amount
                elif "Reserve" in account:
                    home.cell(row = 23, column= column_index_from_string("L")).value = amount
                elif "Growth" in account:
                    home.cell(row = 24, column= column_index_from_string("L")).value = amount

                # refresh date
                home.cell(row = 5, column= column_index_from_string("C")).value = date

                for i in range(22-24):
                    wb.cellStyle(Style.RIGHTALIGN, HomeColor.BANKFILL, 
                                home.cell(row=i, column= column_index_from_string("L")))

            elif "Invest" in account:
                home.cell(row = 6, column= column_index_from_string("B")).value = amount
                home.cell(row = 6, column= column_index_from_string("C")).value = date

            elif "Safe" in account:
                home.cell(row = 7, column= column_index_from_string("B")).value = amount
                home.cell(row = 7, column= column_index_from_string("C")).value = date

            elif "Wallet" in account:
                home.cell(row = 8, column= column_index_from_string("B")).value = amount
                home.cell(row = 8, column= column_index_from_string("C")).value = date

            elif "Gift Card 1" in account:
                home.cell(row = 9, column= column_index_from_string("B")).value = amount
                home.cell(row = 9, column= column_index_from_string("C")).value = date

            elif "Gift Card 2" in account:
                home.cell(row = 10, column= column_index_from_string("B")).value = amount
                home.cell(row = 10, column= column_index_from_string("C")).value = date

            else:
                # need to find a way to get the sum of all "other" (within mysql???)
                home.cell(row = 11, column= column_index_from_string("B")).value = amount
                home.cell(row = 11, column= column_index_from_string("C")).value = date

        #applying styles to all Net Worth cells
        for i in range(5,12):
            wb.cellStyle(Style.RIGHTALIGN, HomeColor.NETWORTHFILL,
                            home.cell(row=i, column=column_index_from_string("B")))
            wb.cellStyle(Style.CENTERALIGN, HomeColor.NETWORTHFILL,
                        home.cell(row=i, column=column_index_from_string("C")))

        self.closeConnection()
        wb.saveFile()


    # ----------------------------------------------databaseTrack sheet----------------------------------
    
    def downloadAccount(self):
        """
        Downloads Account Tracking from database to spreadsheet
        Implemented
        """
        cursor = self.cursor

        wb = ExcelSheet()
        dataTrack = wb.dataTrack
        
        """
        getting last account_id
        """
        lastId = lastRow = 0
        firstId = dataTrack.cell(row = 3, column= column_index_from_string("A")).value
        # if first cell is empty
        if firstId == None:
            lastId = 1000
            lastRow = 3
        else:
            # look for first empty cell
            for i in range(2, dataTrack.max_row):
                nextID = dataTrack.cell(row = i+1, column= column_index_from_string("A")).value

                if nextID == None:
                    lastId = dataTrack.cell(row = i, column= column_index_from_string("A")).value
                    lastRow = i+1
                    
                    break

        """
        getting all account data from database
        """
        query = """select * from account where acct_id > %d""" % (lastId)
        cursor.execute(query)
        result = cursor.fetchall() #array of 4

        """
        downloading all account data to spreadsheet
        """
        for tracked in result:
            id = tracked[0]
            name = tracked[1]
            value = tracked[2]
            day = tracked[3]

            date = day.strftime("%m/%d/%Y")

            dataTrack.cell(row = lastRow, column= column_index_from_string("A")).value = id
            dataTrack.cell(row = lastRow, column= column_index_from_string("B")).value = name
            dataTrack.cell(row = lastRow, column= column_index_from_string("C")).value = value
            dataTrack.cell(row = lastRow, column= column_index_from_string("D")).value = date

            wb.cellStyle(Style.LEFTALIGN, DatabaseTrackColor.ACCOUNTFILL, 
                        dataTrack.cell(row=lastRow, column= column_index_from_string("A")))
            wb.cellStyle(Style.LEFTALIGN, DatabaseTrackColor.ACCOUNTFILL, 
                        dataTrack.cell(row=lastRow, column= column_index_from_string("B")))
            wb.cellStyle(Style.RIGHTALIGN, DatabaseTrackColor.ACCOUNTFILL, 
                        dataTrack.cell(row=lastRow, column= column_index_from_string("C")))
            wb.cellStyle(Style.CENTERALIGN, DatabaseTrackColor.ACCOUNTFILL, 
                        dataTrack.cell(row=lastRow, column= column_index_from_string("D")))
            
            lastRow += 1

        self.closeConnection()
        wb.saveFile()

    def downloadProfit(self):
        """
        Downloads Profit Trackign from database to spreadsheet
        Implemented
        note: cells H1 and I1 are filled in and have borders
        """
        cursor = self.cursor

        wb = ExcelSheet()
        dataTrack = wb.dataTrack

        """
        getting last profit_id
        """
        lastId = lastRow = 0
        firstId = firstId = dataTrack.cell(row = 3, column= column_index_from_string("F")).value
        # if first cell is empty
        if firstId == None:
            lastId = 1000
            lastRow = 3
        else:
            for i in range(2, dataTrack.max_row):
                nextID = dataTrack.cell(row = i+1, column= column_index_from_string("F")).value

                if nextID == None:
                    lastId = dataTrack.cell(row = i, column= column_index_from_string("F")).value
                    lastRow = i+1
                    
                    break

        """
        getting all profit data from database
        """
        query = """select * from profit where profit_id > %d""" % (lastId)
        cursor.execute(query)
        result = cursor.fetchall() #array of 4

        """
        downloading all account data to spreadsheet
        """
        for tracked in result:
            id = tracked[0]
            value = tracked[1]
            day = tracked[2]
            time = tracked[3]

            # print(id, name, value, day)

            date = day.strftime("%m/%d/%Y")

            dataTrack.cell(row = lastRow, column= column_index_from_string("F")).value = id
            dataTrack.cell(row = lastRow, column= column_index_from_string("G")).value = value
            dataTrack.cell(row = lastRow, column= column_index_from_string("H")).value = date
            dataTrack.cell(row = lastRow, column= column_index_from_string("I")).value = time

            wb.cellStyle(Style.LEFTALIGN, DatabaseTrackColor.PROFITFILL, 
                        dataTrack.cell(row=lastRow, column= column_index_from_string("F")))
            wb.cellStyle(Style.RIGHTALIGN, DatabaseTrackColor.PROFITFILL, 
                        dataTrack.cell(row=lastRow, column= column_index_from_string("G")))
            wb.cellStyle(Style.CENTERALIGN, DatabaseTrackColor.PROFITFILL, 
                        dataTrack.cell(row=lastRow, column= column_index_from_string("H")))
            wb.cellStyle(Style.CENTERALIGN, DatabaseTrackColor.PROFITFILL, 
                        dataTrack.cell(row=lastRow, column= column_index_from_string("I")))
            
            lastRow += 1

        self.closeConnection()
        wb.saveFile()

    def newLogin(self):
        """
        records all new logins on database and spreadsheet
        Implemented
        """
        # adding a new sign in to database
        cursor = self.cursor
        query = "call newLogin"
        cursor.execute(query)


        wb = ExcelSheet()
        dataTrack = wb.dataTrack

        """
        getting last account_login
        """
        lastId = lastRow = 0
        firstId = dataTrack.cell(row = 3, column= column_index_from_string("K")).value
        # if first cell is empty
        if firstId == None:
            lastId = 1000
            lastRow = 3
        else:
            # look for first empty cell
            for i in range(2, dataTrack.max_row):
                nextID = dataTrack.cell(row = i+1, column= column_index_from_string("K")).value

                if nextID == None:
                    lastId = dataTrack.cell(row = i, column= column_index_from_string("K")).value
                    lastRow = i+1
                    break

            
        """
        getting all login data from database
        """
        query = """select * from login_track where login_id > %d""" % (lastId)
        cursor.execute(query)
        result = cursor.fetchall() #array of 3

        """
        downloading all login data to spreadsheet
        """
        for tracked in result:
            index = tracked[0]
            day = tracked[1] 
            sec = tracked[2] 

            date = day.strftime("%m/%d/%Y")
 
            dataTrack.cell(row = lastRow, column= column_index_from_string("K")).value = index
            dataTrack.cell(row = lastRow, column= column_index_from_string("L")).value = date
            dataTrack.cell(row = lastRow, column= column_index_from_string("M")).value = sec

            wb.cellStyle(Style.LEFTALIGN, DatabaseTrackColor.LOGINFILL, 
                        dataTrack.cell(row = lastRow, column = column_index_from_string("K")))
            wb.cellStyle(Style.CENTERALIGN, DatabaseTrackColor.LOGINFILL, 
                        dataTrack.cell(row = lastRow, column = column_index_from_string("L")))
            wb.cellStyle(Style.RIGHTALIGN, DatabaseTrackColor.LOGINFILL, 
                        dataTrack.cell(row = lastRow, column = column_index_from_string("M")))

            lastRow += 1
        
        self.closeConnection()
        wb.saveFile()

    def downloadNetWorth(self):
        """
        Downloads Net Worth Tracking from database to spreadsheet
        Implemented
        """
        cursor = self.cursor

        wb = ExcelSheet()
        dataTrack = wb.dataTrack
        
        """
        getting last networth_id
        """
        lastId = lastRow = 0
        firstId = dataTrack.cell(row = 3, column= column_index_from_string("O")).value
        # if first cell is empty
        if firstId == None:
            lastId = 1000
            lastRow = 3
        else:
            # look for first empty cell
            for i in range(2, dataTrack.max_row):
                nextID = dataTrack.cell(row = i+1, column= column_index_from_string("O")).value

                if nextID == None:
                    lastId = dataTrack.cell(row = i, column= column_index_from_string("O")).value
                    lastRow = i+1
                    
                    break

        """
        getting all account data from database
        """
        query = """select * from netWorth where worth_id > %d""" % (lastId)
        cursor.execute(query)
        result = cursor.fetchall() #array of 4

        """
        downloading all account data to spreadsheet
        """
        for tracked in result:
            id = tracked[0]
            value = tracked[1]
            day = tracked[2]
            time = tracked[3]

            date = day.strftime("%m/%d/%Y")

            dataTrack.cell(row = lastRow, column= column_index_from_string("O")).value = id
            dataTrack.cell(row = lastRow, column= column_index_from_string("P")).value = value
            dataTrack.cell(row = lastRow, column= column_index_from_string("Q")).value = date
            dataTrack.cell(row = lastRow, column= column_index_from_string("R")).value = time

            wb.cellStyle(Style.LEFTALIGN, DatabaseTrackColor.NETWORTHFILL, 
                        dataTrack.cell(row=lastRow, column= column_index_from_string("O")))
            wb.cellStyle(Style.RIGHTALIGN, DatabaseTrackColor.NETWORTHFILL, 
                        dataTrack.cell(row=lastRow, column= column_index_from_string("P")))
            wb.cellStyle(Style.CENTERALIGN, DatabaseTrackColor.NETWORTHFILL, 
                        dataTrack.cell(row=lastRow, column= column_index_from_string("Q")))
            wb.cellStyle(Style.CENTERALIGN, DatabaseTrackColor.NETWORTHFILL, 
                        dataTrack.cell(row=lastRow, column= column_index_from_string("R")))
            
            lastRow += 1

        self.closeConnection()
        wb.saveFile()


    # ----------------------------------------------Paychecks sheet----------------------------------
    def downloadPaycheck(self):
        """
        Downloads Paycheck Trackign from database to spreadsheet
        Implemented
        """
        cursor = self.cursor

        wb = ExcelSheet()
        paycheck = wb.paychecks

        """
        getting last paycheck_id
        """
        lastId = lastRow = 0
        firstId = firstId = paycheck.cell(row = 6, column= column_index_from_string("A")).value
        # if first cell is empty
        if firstId == None:
            lastId = 1000
            lastRow = 6
        else:
            for i in range(5, paycheck.max_row):
                nextID = paycheck.cell(row = i+1, column= column_index_from_string("A")).value

                if nextID == None:
                    lastId = paycheck.cell(row = i, column= column_index_from_string("A")).value
                    lastRow = i+1
                    
                    break

        """
        getting all profit data from database
        """
        query = """select * from paycheck where check_id > %d""" % (lastId)
        cursor.execute(query)
        result = cursor.fetchall() #array of 12 (2 not counted (income_id))

        """
        downloading all account data to spreadsheet
        """
        for tracked in result:
            id= tracked[0]
            company = tracked[2]
            hours = tracked[3]
            start = tracked[4]
            end = tracked[5]
            pay = tracked[6]
            gross = tracked[7]
            federal = tracked[8]
            state = tracked[9]
            city = tracked[10]
            final = tracked[11]

            startD = start.strftime("%m/%d/%Y")
            endD = end.strftime("%m/%d/%Y")
            payD = pay.strftime("%m/%d/%Y")

            paycheck.cell(row = lastRow, column= column_index_from_string("A")).value = id
            paycheck.cell(row = lastRow, column= column_index_from_string("B")).value = company
            paycheck.cell(row = lastRow, column= column_index_from_string("F")).value = hours
            paycheck.cell(row = lastRow, column= column_index_from_string("C")).value = startD
            paycheck.cell(row = lastRow, column= column_index_from_string("D")).value = endD
            paycheck.cell(row = lastRow, column= column_index_from_string("E")).value = payD
            paycheck.cell(row = lastRow, column= column_index_from_string("G")).value = gross
            paycheck.cell(row = lastRow, column= column_index_from_string("H")).value = federal
            paycheck.cell(row = lastRow, column= column_index_from_string("I")).value = state
            paycheck.cell(row = lastRow, column= column_index_from_string("J")).value = city
            paycheck.cell(row = lastRow, column= column_index_from_string("K")).value = final


            wb.cellStyle(Style.LEFTALIGN, PaycheckColor.PAYCHECKFILL, 
                        paycheck.cell(row=lastRow, column= column_index_from_string("A")))
            wb.cellStyle(Style.LEFTALIGN, PaycheckColor.PAYCHECKFILL, 
                        paycheck.cell(row=lastRow, column= column_index_from_string("B")))
            wb.cellStyle(Style.CENTERALIGN, PaycheckColor.PAYCHECKFILL, 
                        paycheck.cell(row=lastRow, column= column_index_from_string("F")))
            wb.cellStyle(Style.CENTERALIGN, PaycheckColor.PAYCHECKFILL, 
                        paycheck.cell(row=lastRow, column= column_index_from_string("C")))
            wb.cellStyle(Style.CENTERALIGN, PaycheckColor.PAYCHECKFILL, 
                        paycheck.cell(row=lastRow, column= column_index_from_string("D")))
            wb.cellStyle(Style.CENTERALIGN, PaycheckColor.PAYCHECKFILL, 
                        paycheck.cell(row=lastRow, column= column_index_from_string("E")))
            wb.cellStyle(Style.RIGHTALIGN, PaycheckColor.PAYCHECKFILL, 
                        paycheck.cell(row=lastRow, column= column_index_from_string("G")))
            wb.cellStyle(Style.RIGHTALIGN, PaycheckColor.PAYCHECKFILL, 
                        paycheck.cell(row=lastRow, column= column_index_from_string("H")))
            wb.cellStyle(Style.RIGHTALIGN, PaycheckColor.PAYCHECKFILL, 
                        paycheck.cell(row=lastRow, column= column_index_from_string("I")))
            wb.cellStyle(Style.RIGHTALIGN, PaycheckColor.PAYCHECKFILL, 
                        paycheck.cell(row=lastRow, column= column_index_from_string("J")))
            wb.cellStyle(Style.RIGHTALIGN, PaycheckColor.PAYCHECKFILL, 
                        paycheck.cell(row=lastRow, column= column_index_from_string("K")))
            
            lastRow += 1

        self.closeConnection()
        wb.saveFile()


    # ----------------------------------------------Month sheet----------------------------------
    def downloadAllExpenses(self):
        """
        downloads all expense data from database to spreadsheet
        only uploads to the month sheets, not the HOME sheet
        not working
        """
        cursor = self.cursor

        wb = ExcelSheet()
        nextMonth = wb.febuary
        
        """
        getting all expenses from database
        """
        query = "select * from expense"
        cursor.execute(query)
        result = cursor.fetchall() # array of 6 (not using first)

        # print(result)
        """
        downloading all expense data to spreadsheets
        """
        i = 5
        for tracked in result:
            amount = tracked[1]
            item = tracked[2]
            party = tracked[3]
            day = tracked[4]
            type = tracked[5]

            date = day.strftime("%m/%d/%Y")
            month = self.checkMonth(date, wb)
            if month != nextMonth:
                i = 5
            nextMonth = month
            print(month)

            month.cell(row = i, column= column_index_from_string("G")).value = item
            month.cell(row = i, column= column_index_from_string("H")).value = party
            month.cell(row = i, column= column_index_from_string("I")).value = date
            month.cell(row = i, column= column_index_from_string("J")).value = amount
            month.cell(row = i, column= column_index_from_string("K")).value = type

            wb.cellStyle(Style.LEFTALIGN, MonthColor.EXPENSESFILL, 
                        month.cell(row=i, column= column_index_from_string("G")))
            wb.cellStyle(Style.CENTERALIGN, MonthColor.EXPENSESFILL, 
                        month.cell(row=i, column= column_index_from_string("H")))
            wb.cellStyle(Style.CENTERALIGN, MonthColor.EXPENSESFILL, 
                        month.cell(row=i, column= column_index_from_string("I")))
            wb.cellStyle(Style.RIGHTALIGN, MonthColor.EXPENSESFILL, 
                        month.cell(row=i, column= column_index_from_string("J")))
            wb.cellStyle(Style.CENTERALIGN, MonthColor.EXPENSESFILL, 
                        month.cell(row=i, column= column_index_from_string("K")))
            
            i += 1

        self.closeConnection
        wb.saveFile()


    def downloadAllIncome(self):
        """
        downloads all expense data from database to spreadsheet
        only uploads to the month sheets, not the HOME sheet
        not working
        """
        cursor = self.cursor

        wb = ExcelSheet()
        nextMonth = wb.febuary
        
        """
        getting all income from database
        """
        query = "select * from income"
        cursor.execute(query)
        result = cursor.fetchall() # array of 6 (not using first)

        # print(result)
        """
        downloading all expense data to spreadsheets
        """
        i = 5
        for tracked in result:
            amount = tracked[1]
            item = tracked[2]
            source = tracked[3]
            day = tracked[4]
            type = tracked[5]

            date = day.strftime("%m/%d/%Y")
            month = self.checkMonth(date, wb)
            if month != nextMonth:
                i = 5
            nextMonth = month
            print(month)

            month.cell(row = i, column= column_index_from_string("A")).value = item
            month.cell(row = i, column= column_index_from_string("B")).value = source
            month.cell(row = i, column= column_index_from_string("C")).value = date
            month.cell(row = i, column= column_index_from_string("D")).value = amount
            month.cell(row = i, column= column_index_from_string("E")).value = type

            wb.cellStyle(Style.LEFTALIGN, MonthColor.INCOMEFILL, 
                        month.cell(row=i, column= column_index_from_string("A")))
            wb.cellStyle(Style.CENTERALIGN, MonthColor.INCOMEFILL, 
                        month.cell(row=i, column= column_index_from_string("B")))
            wb.cellStyle(Style.CENTERALIGN, MonthColor.INCOMEFILL, 
                        month.cell(row=i, column= column_index_from_string("C")))
            wb.cellStyle(Style.RIGHTALIGN, MonthColor.INCOMEFILL, 
                        month.cell(row=i, column= column_index_from_string("D")))
            wb.cellStyle(Style.CENTERALIGN, MonthColor.INCOMEFILL, 
                        month.cell(row=i, column= column_index_from_string("E")))
            
            i += 1

        self.closeConnection
        wb.saveFile()


    # ----------------------------------------------Helper methods-------------------------------------
    def lastLogin(self):
        """
        Returns the date of the last login on system
        Implemented
        """
        wb = ExcelSheet()
        dataTrack = wb.dataTrack

        """
        Getting Last Login Date
        """
        lastDate = None
        firstDate = dataTrack.cell(row = 3, column= column_index_from_string("L")).value
        # if first cell is empty
        if firstDate == None:
            lastDate = firstDate
        else:
            # look for first empty cell
            for i in range(3, dataTrack.max_row):
                nextID = dataTrack.cell(row = i+1, column= column_index_from_string("L")).value

                if nextID == None:
                    lastDate = dataTrack.cell(row = i, column= column_index_from_string("L")).value
                    break
        
        return lastDate.strftime("%m/%d/%Y")

    def checkMonth(self, date, wb):
        """
        Checking the month of a given date
        Used for income and expenses
        Implemented
        """

        # taking month substring out of date
        month = date[0:2:1]

        if month == "01":
            return wb.january
        elif month == "02":
            return wb.febuary
        elif month == "03":
            return wb.march
        elif month == "04":
            return wb.april
        elif month == "05":
            return wb.may
        elif month == "06":
            return wb.june
        elif month == "07":
            return wb.july
        elif month == "08":
            return wb.august
        elif month == "09":
            return wb.september
        elif month == "10":
            return wb.october
        elif month == "11":
            return wb.november
        elif month == "12":
            return wb.december
        
class TestMethods():

    # showing each value
    def netWorth():
        index = 1
        for row in range(5, 12):
            print(index, ") ", master.cell(row = row, column = 1).value,
                "\t\t\t$", master.cell(row = row, column = 2).value)
            index= index +1

    def income():
        index = 1
        for row in range(5, 17):
            print(index,") ", master.cell(row = row, column = 9).value,
                "\t", master.cell(row = row, column = 10).value,
                "\t$", master.cell(row=row, column = 11).value)
            index = index +1

    def purchase():
        index = 1
        for row in range(5, 30):
            print(index,") ", master.cell(row = row, column = 13).value,
                "\t", master.cell(row = row, column = 14).value,
                "\t$", master.cell(row=row, column = 15).value)
            index = index +1


    # calculating totals
    def totalWorth():
        worth = 0.00
        for row in range(5, 12):
            if (master.cell(row=row, column = 2).value != None):
                worth = worth + (master.cell(row = row, column = 2).value)
        print("$", "{:.2f}".format(worth))

    def totalRecieve():
        worth = 0.00
        for row in range(5, 17):
            if (master.cell(row=row, column = 11).value != None):
                worth = worth + (master.cell(row = row, column = 11).value)
        print("$", "{:.2f}".format(worth))

    def totalBought():
        worth = 0.00
        for row in range(5, 30):
            if (master.cell(row=row, column = 15).value != None):
                worth = worth +master.cell(row = row, column = 15).value
        print("$", "{:.2f}".format(worth))

    # creates a new sheet for the month
    def newMonthSheet(month):
        print(month)
        wb.create_sheet('test')
        wb.save
        print("success")

    # searches for the next month that needs a sheet
    def newMonth():
        if ('January' not in wb.sheetnames):
            newMonthSheet('January')
        elif ('Febuary' not in wb.sheetnames):
            newMonthSheet('Febuary')
        elif ('March' not in wb.sheetnames):
            newMonthSheet('March')
        elif ('April' not in wb.sheetnames):
            newMonthSheet('April')
        elif ('May' not in wb.sheetnames):
            newMonthSheet('May')
        elif ('June' not in wb.sheetnames):
            newMonthSheet('June')
        elif ('July' not in wb.sheetnames):
            newMonthSheet('July')
        elif ('August' not in wb.sheetnames):
            newMonthSheet('August')
        elif ('September' not in wb.sheetnames):
            newMonthSheet('September')
        elif ('October' not in wb.sheetnames):
            newMonthSheet('October')
        elif ('November' not in wb.sheetnames):
            newMonthSheet('November')
        elif ('December' not in wb.sheetnames):
            newMonthSheet('December')
            
    def updatePNC():
        #updating last updated for PNC
        string = "PNC Bank acct (as of "
        string2 = time.strftime("%d/%m/%Y")
        string = string + string2 + ")"
        home['A5'] = string
        print(string)

    def updateMorgan():
        string = "Morgan Stanley (as of "
        string2 = time.strftime("%d/%m/%Y")
        string = string + string2 + ")"
        master['A6'] = string
        print(string)

    def newIncome(income):
        column = 11 #K column
        for i in range (5,17):
            if master.cell(i, column).value == None:
                master.cell(i, column).value == income
                wb.save(filename)
                break

    def newPurchase(purchase):
        column = 15 #O column
        for i in range (5,30):
            if master.cell(i, column).value == None:
                master.cell(i, column).value == purchase
                wb.save(filename)
                break            


# with sshtunnel.SSHTunnelForwarder(('172.93.207.7'),
#       ssh_username = 'pi', ssh_password = 'irohs_army51!', remote_bind_address=()
#    ) as tunnel:
#    connection = mysql.connect(
#       user = 'project',
#       passwd = 'trackerPass',
#       host = '192.168.1.38', port = tunnel.local_bind_port,
#       db = 'financial',
#    )
# print(sshtunnel.check_address(("172.93.207.7", 22)))

# local connection (WORKS!!!)
print("Begin")
db = mysql.connect(
            host = "192.168.1.38",
            database = "financial",
            user = "project",
            password = "trackerPass"
            )

test = DatabaseQuery(db)
# # test.testDBQuery()
# # test.newLogin()
# test.downloadAccount()
# test.downloadProfit()
# test.downloadNetWorth()
test.downloadPaycheck()
# test.updateSubscription()
# test.updateDesiredPurchase()
# test.updateForSale()
# # print(test.lastLogin())
# test.updateNetWorth()

# # print(test.checkMonth(test.lastLogin()))
# # test.downloadAllExpenses()
# # test.downloadAllIncome()



test.closeConnection()
print("End")

exit()
