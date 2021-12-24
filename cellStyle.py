from openpyxl.styles import Alignment, Color, PatternFill, Border
from openpyxl.styles.borders import Side


class Style():
    """
    Implemented
    """
    ALLBORDER = Border(left=Side(style='thin'),
                        right= Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
    LEFTALIGN = Alignment(horizontal='left')
    RIGHTALIGN = Alignment(horizontal='right')
    CENTERALIGN = Alignment(horizontal='center')

class HomeColor():
    """
    Implemented
    """
    NETWORTHFILL = PatternFill(patternType='solid', 
                            fill_type= 'solid', 
                            fgColor= Color('00FFFF'))
    SUBSCRIPTIONFILL = PatternFill(patternType='solid', 
                            fill_type= 'solid', 
                            fgColor= Color('93C47D'))
    DESIREDPURFILL = PatternFill(patternType='solid', 
                            fill_type= 'solid', 
                            fgColor= Color('00FF00'))
    FORSALEFILL = PatternFill(patternType='solid', 
                            fill_type= 'solid', 
                            fgColor= Color('6D9EEB'))
    BANKFILL = PatternFill(patternType='solid', 
                            fill_type= 'solid', 
                            fgColor= Color('F4B084'))
    INCOMEFILL = PatternFill(patternType='solid', 
                            fill_type= 'solid', 
                            fgColor= Color('6FA8DC'))
    EXPENSESFILL = PatternFill(patternType='solid', 
                            fill_type= 'solid', 
                            fgColor= Color('A9D08E'))

class PaycheckColor():
    """
    Implemented
    """
    PAYCHECKFILL = PatternFill(patternType='solid', 
                            fill_type= 'solid', 
                            fgColor= Color('F4B084'))

class DatabaseTrackColor():
    """
    Implemented
    """
    LOGINFILL = PatternFill(patternType='solid', 
                            fill_type= 'solid', 
                            fgColor= Color('A9D08E'))
    PROFITFILL = PatternFill(patternType='solid', 
                            fill_type= 'solid', 
                            fgColor= Color('8EA9DB'))
    ACCOUNTFILL = PatternFill(patternType='solid', 
                            fill_type= 'solid', 
                            fgColor= Color('ED7D31'))
    NETWORTHFILL = PatternFill(patternType='solid', 
                            fill_type= 'solid', 
                            fgColor= Color('FFD966'))

class MonthColor():
    """
    Not Implemented
    """
    INCOMEFILL = PatternFill(patternType='solid', 
                            fill_type= 'solid', 
                            fgColor= Color('ED7D31'))
    EXPENSESFILL = PatternFill(patternType='solid', 
                            fill_type= 'solid', 
                            fgColor= Color('5B9BD5'))